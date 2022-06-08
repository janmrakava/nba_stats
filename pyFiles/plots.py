# %% [markdown]
# # Plots from scraped data

# %%
%pip install matplotlib
%pip install openpyxl

# %% [markdown]
# ### Plot for average points per player for all seasons

# %%
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt
from openpyxl import load_workbook
%matplotlib inline

dict_df = pd.read_excel('excelFiles/groupByTeams.xlsx', sheet_name=None)
wb = openpyxl.load_workbook("excelFiles/groupByTeams.xlsx")
sheet_names = (wb.sheetnames)
dfA = pd.DataFrame(columns=['SEASON', 'PTS'])
seasons = []
games = []
for season in sheet_names:
    seasonList = dict_df.get(season)
    gp = seasonList.loc[seasonList['TEAM'] == 'GSW','PTS']
    games.append(gp.values[0])
    seasons.append(season)

dataF = pd.DataFrame(seasons, columns=['SEASON'])
dataF['PTS'] = games
dataF['SEASON'] = seasons
fig = dataF.plot(x='SEASON', y = 'PTS', figsize=(25,10), title="Průměrný počet bodů na hráče v průběhu 25 sezon týmu Golden State Warriors").get_figure()
plt.xticks(range(0,len(seasons)), seasons)
fig.savefig('plots/avgPoints.png')

# %% [markdown]
# ### Plot for average number of wins of all teams for all seasons

# %%
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt
from openpyxl import load_workbook
%matplotlib inline
%store -r teamsAll

dict_df = pd.read_excel('excelFiles/teams.xlsx')
wb = openpyxl.load_workbook("excelFiles/teams.xlsx")
sheet_names = (wb.sheetnames)
dfA = pd.DataFrame(columns=['TEAM', 'W'])
teams = []
wins = []
for team in teamsAll:
    w = seasonList[seasonList['TEAM'] == team]['W']
    wins.append(w.values[0])
    teams.append(team)
    
dataF = pd.DataFrame(teams, columns=['TEAMS'])
dataF['WINS'] = wins
dataF['TEAMS'] = teams
dataF.sort_values(by=['WINS'], ascending=False, inplace=True)
fig = dataF.plot(kind='bar', x='TEAMS', y = 'WINS', figsize=(20,10), title="Průměrný počet výher týmů za 25 sezon").get_figure()
plt.xticks(range(0,len(teams)), teams)
fig.savefig('plots/allTeamPoints.png')

# %% [markdown]
# ### Point development of player by all seasons

# %%
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt
from openpyxl import load_workbook
%matplotlib inline

def points_by_season(player): 
    dict_df = pd.read_excel('excelFiles/stats.xlsx', sheet_name=None)
    wb = openpyxl.load_workbook("excelFiles/stats.xlsx")
    sheet_names = (wb.sheetnames)
    dfA = pd.DataFrame(columns=['SEASON', 'PTS'])
    seasons = []
    points = []
    for season in sheet_names:
        seasonList = dict_df.get(season)
        pts = seasonList.loc[seasonList['PLAYER'] == player]['PTS']
        if pts.empty:
            points.append('0')
        else:    
            points.append(pts.values[0])
        seasons.append(season)
    dataF = pd.DataFrame(seasons, columns=['SEASON'])
    dataF['PTS'] = points
    dataF['SEASON'] = seasons
    dataF['PTS'] = dataF['PTS'].astype(int)

    fig = dataF.plot(x='SEASON', y = 'PTS', figsize=(25,10), title=f"Počet bodů v sezoně hráče {player} od sezony 1996-97").get_figure()
    plt.xticks(range(0,len(seasons)), seasons)
    fig.savefig(f'plots/points_{player}.png')
points_by_season("Giannis Antetokounmpo")

# %% [markdown]
# ### Write all plots into own .xlsx file

# %%
import openpyxl

wrkb = openpyxl.Workbook()
  
ws = wrkb.worksheets[0]
img = openpyxl.drawing.image.Image('plots/avgPoints.png')
img2 = openpyxl.drawing.image.Image('plots/allTeamPoints.png')

img.anchor = 'A1'
img2.anchor = 'A40'
ws.add_image(img)
ws.add_image(img2)
wrkb.save('excelFiles/plots.xlsx')
print('All plots saved')


