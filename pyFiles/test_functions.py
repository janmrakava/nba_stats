# %% [markdown]
# # Testing functions - not used in project

# %% [markdown]
# ## Funkce pro vrácení sumy bodů jednotlivých týmů za všechny sezóny 
# 
# TODO: seřadit sestupně pro zápis do xlsx

# %%
def sum_team_points_all_seasons():
    soucet = 0    
    for team in teamsAll:   
         for index in sheet_names:   
            seasonList = dict_df.get(index)                 
            soucet += seasonList[seasonList['TEAM']==team].sum()['PTS']
         print(f"Součet pro tým {team} je {soucet} bodů.")
         soucet = 0
         #TODO: seřadit dle počtu sestupně?
sum_team_points_all_seasons()

# %% [markdown]
# ## Funkce pro výpočet procentuálních výher a proher za sezonu
# 
# TODO: obecně pro všechny hráče i do listu all_players?
# 

# %%
def number_W_L_per_season(playerName):
    for index in sheet_names:
        seasonList = dict_df.get(index)
        print(index)
        vyher = seasonList.loc[seasonList['PLAYER'] == playerName, 'W'].sum()
        proher = seasonList.loc[seasonList['PLAYER'] == playerName, 'L'].sum()      
        pocetZapasu = vyher+proher
        if (pocetZapasu > 0):
            vyher = round((vyher/pocetZapasu)*100,2);
            proher = round((proher/pocetZapasu)*100,2);
            print(f"Hráč {playerName} za sezonu {index} vyhrál {vyher}% zápasů.")
            print(f"Hráč {playerName} za sezonu {index} prohrál {proher}% zápasů.")
        else:
            print('Hráč neodehrál žádný zápas')
number_W_L_per_season('Kevin Durant')

# %% [markdown]
# ## Funkce pro výpočet procentuálních výher a proher za celou dobu
# 
# Kontrolní fce, jestli fce number_W_L_all_time() funguje, jak má. 

# %%
def number_W_L_all_time(playerName):
    vyher = 0
    proher = 0
    for index in sheet_names:
        seasonList = dict_df.get(index)        
        vyher += seasonList.loc[seasonList['PLAYER'] == playerName, 'W'].sum()
        proher += seasonList.loc[seasonList['PLAYER'] == playerName, 'L'].sum()   
    pocetZapasu = vyher+proher
    vyher = round((vyher/pocetZapasu)*100,2)
    proher = round((proher/pocetZapasu)*100,2)
    print(f"{playerName} odehrál {pocetZapasu} zápasů a z toho jich {vyher}% vyhrál, {proher}% prohrál")          
number_W_L_all_time('Michael Jordan')

# %% [markdown]
# ## Funkce pro spočítání úspěšnosti field goals za celou dobu
# 
# Kontrolní fce, zda-li fce succ_field_goals() funguje tak, jak má.

# %%
def succ_field_goals(playerName):
    celkovyPocet = 0
    pocetUspesnych = 0
    for index in sheet_names:
        seasonList = dict_df.get(index) 
        celkovyPocet += seasonList.loc[seasonList['PLAYER'] == playerName, 'FGA'].sum()  
        pocetUspesnych += seasonList.loc[seasonList['PLAYER'] == playerName, 'FGM'].sum()
    uspesnost = round((pocetUspesnych/celkovyPocet)*100,2)
    print(f"Hráč {playerName} se za svoji kariéru v NBA pokusil celkem o {celkovyPocet} hodů na koš, bez trestných hodů, z toho bylo {pocetUspesnych} úspěšných hodů, {celkovyPocet-pocetUspesnych} neúspěšných hodů, procentuálně to vychází na {uspesnost}% úspěšnost")
succ_field_goals('Derrick Rose')   

# %% [markdown]
# ## Úspěšnost trojek pro hráče za celou dobu
# 
# Kontrolní fce, zda-li fce succ_three_goals() funguje tak, jak má.
# 
# - Problém by mohl být, že množina pokusů o hození trojky, je podmnožinou field goals

# %%
def succ_three_goals(playerName):
    celkovyPocet = 0
    pocetUspesnych = 0
    for index in sheet_names:
        seasonList = dict_df.get(index) 
        celkovyPocet += seasonList.loc[seasonList['PLAYER'] == playerName, '3PA'].sum()  
        pocetUspesnych += seasonList.loc[seasonList['PLAYER'] == playerName, '3PM'].sum()
    uspesnost = round((pocetUspesnych/celkovyPocet)*100,2)
    print(f"Hráč {playerName} se za svoji kariéru v NBA pokusil celkem o {celkovyPocet} tříbodových hodů na koš, z toho bylo {pocetUspesnych} úspěšných hodů, {celkovyPocet-pocetUspesnych} neúspěšných hodů, procentuálně to vychází na {uspesnost}% úspěšnost")
succ_three_goals('Didi Louzada')   

# %% [markdown]
# ## Úspěšnost trestných hodů za celou dobu
# 
# Kontrolní fce, zda-li fce succ_free_throws() funguje tak, jak má.

# %%
def succ_free_throws(playerName):
    celkovyPocet = 0
    pocetUspesnych = 0
    for index in sheet_names:
        seasonList = dict_df.get(index) 
        celkovyPocet += seasonList.loc[seasonList['PLAYER'] == playerName, 'FTA'].sum()  
        pocetUspesnych += seasonList.loc[seasonList['PLAYER'] == playerName, 'FTM'].sum()
    uspesnost = round((pocetUspesnych/celkovyPocet)*100,2)
    print(f"Hráč {playerName} se za svoji kariéru v NBA pokusil celkem o {celkovyPocet} trestných hodů na koš, z toho bylo {pocetUspesnych} úspěšných hodů, {celkovyPocet-pocetUspesnych} neúspěšných hodů, procentuálně to vychází na {uspesnost}% úspěšnost")
succ_free_throws('Kevin Durant')   

# %% [markdown]
# ## Funkce na výpočet průměrného počtu asistencí za celou dobu
# 
# Kontrolní fce, jestli fce avg_assists() funguje tak, jak má. 

# %%
def avg_assists(playerName):
    celkovyPocet = 0
    pocetZapasu = 0
    for index in sheet_names:
        seasonList = dict_df.get(index) 
        celkovyPocet += seasonList.loc[seasonList['PLAYER'] == playerName, 'AST'].sum()  
        pocetZapasu += seasonList.loc[seasonList['PLAYER'] == playerName, 'GP'].sum()
    prumer = round((celkovyPocet/pocetZapasu),2)
    print(f"Hráč {playerName} má průměrný počet asistencí {prumer} za zápas")
avg_assists('LeBron James')   

# %% [markdown]
# ## Funkce na výpočet průměrného počtu doskoků za celou dobu na zápas
# 
# Kontrolní fce, jestli fce avg_reb() funguje tak, jak má.

# %%
def avg_rebounds_cont(playerName):
    celkovyPocet = 0
    pocetZapasu = 0
    for index in sheet_names:
        seasonList = dict_df.get(index) 
        celkovyPocet += seasonList.loc[seasonList['PLAYER'] == playerName, 'REB'].sum()  
        pocetZapasu += seasonList.loc[seasonList['PLAYER'] == playerName, 'GP'].sum()
    prumer = round((celkovyPocet/pocetZapasu),2)
    print(f"Hráč {playerName} má průměrný počet doskoků {prumer} za zápas")
avg_rebounds('Derrick Rose')   

# %% [markdown]
# ## Funkce pro výpis všech unikátních hráčů za 10 sezon v NBA
# 
# Asi možné smazat, protože je to již zmíněno nahoře 

# %%

from openpyxl import load_workbook
import pandas as pd
import openpyxl

dict_df = pd.read_excel('stats.xlsx', sheet_name=None)
wb = openpyxl.load_workbook("stats.xlsx")
sheet_names = (wb.sheetnames)

teamsAll = []
for index in sheet_names: 
    row_sum = dict_df[index]['TEAM']
    for indexy in row_sum:
        if indexy not in teamsAll:
            teamsAll.append(indexy)
            
#Dataframe pro týmy, pak se nakonec zapíše do excelu
df_teams = pd.DataFrame (teamsAll, columns = ['TEAM'])   
print(df_teams)        
def all_players():
    allPlayers = []
    for index in sheet_names:
        seasonList = dict_df.get(index)
        players = seasonList['PLAYER']
        for player in players:
            if player not in allPlayers:
                allPlayers.append(player)
    #book = load_workbook('stats.xlsx', 'rb')
    #writer = pd.ExcelWriter('stats.xlsx')
    #writer.book = book
    df = pd.DataFrame (allPlayers, columns = ['PLAYER']) 
    df_all = pd.DataFrame (allPlayers, columns = ['PLAYER'])       
    print(df) #1392 hráčů, bez duplicit
    
    #df.to_excel(writer, sheet_name="all_players", index=False)
all_players()

# %% [markdown]
# ## Funkce pro výpočet průměrných bodů za sezonu pro hráče
# 
# Kontrolní fce, zda-li fce sum_point_all_players() funguje tak, jak má. 

# %%

def sum_for_player(playerName):
    prumer = 0
    for index in sheet_names:   
        seasonList = dict_df.get(index)  
        print(index)
        soucet = seasonList.loc[seasonList['PLAYER'] == playerName, 'PTS'].sum()
        pocet = seasonList.loc[seasonList['PLAYER'] == playerName, 'GP'].sum()       
        py_int_soucet = soucet.item() ##Takhle to dělám protože je to numpy type a takhle se převádí. Bez toho výpočet vrací hodnotu inf
        py_int_pocet = pocet.item() 
        if (py_int_pocet > 0):            
            prumer = py_int_soucet/py_int_pocet 
            print(f"Průměrný počet bodů za {index} pro hráče {playerName} je {prumer}")              
        else: 
            print('Hráč neodehrál žádný zápas v sezóně')                    
sum_for_player('Kevin Durant')

    

# %% [markdown]
# ## Funkce pro výpočet celkového průměru za všechny sezony pro hráče
# 
# Kontrolní fce, zda-li fce sum_point_all_players() funguje tak, jak má. 

# %%
def sum_for_player_all_season(playerName):
    soucet = 0
    pocet = 0
    for index in sheet_names:
        seasonList = dict_df.get(index)
        soucet += seasonList.loc[seasonList['PLAYER'] == playerName, 'PTS'].sum()
        pocet += seasonList.loc[seasonList['PLAYER'] == playerName, 'GP'].sum()
    prumer = soucet.item()/pocet.item()
    print(f"Průměrný počet bodů za posledních 10 sezon {playerName} je {prumer}")     
sum_for_player_all_season('Michael Jordan')
        

# %% [markdown]
# # Kontrolní fce pro průměrný počet steals za celou dobu pro hráče

# %%
def avgSTL_for_player_all_season(playerName):
    soucet = 0
    pocet = 0
    for index in sheet_names:
        seasonList = dict_df.get(index)
        soucet += seasonList.loc[seasonList['PLAYER'] == playerName, 'STL'].sum()
        pocet += seasonList.loc[seasonList['PLAYER'] == playerName, 'GP'].sum()
    prumer = soucet.item()/pocet.item()
    print(f"Průměrný počet STL za posledních 10 sezon {playerName} je {prumer}")     
avgSTL_for_player_all_season('Michael Jordan')
        

# %% [markdown]
# # Kontrolní fce pro průměrná počet počet TO pro hráče

# %%
def avgTO_for_player_all_season(playerName):
    soucet = 0
    pocet = 0
    for index in sheet_names:
        seasonList = dict_df.get(index)
        soucet += seasonList.loc[seasonList['PLAYER'] == playerName, 'TOV'].sum()
        pocet += seasonList.loc[seasonList['PLAYER'] == playerName, 'GP'].sum()
    prumer = soucet.item()/pocet.item()
    print(f"Průměrný počet TO za posledních 10 sezon {playerName} je {prumer}")     
avgTO_for_player_all_season('Michael Jordan')
        

# %% [markdown]
# # Kontrolní funkce pro avg_blocks()

# %%
def avgBLK_for_player_all_season(playerName):
    soucet = 0
    pocet = 0
    for index in sheet_names:
        seasonList = dict_df.get(index)
        soucet += seasonList.loc[seasonList['PLAYER'] == playerName, 'BLK'].sum()
        pocet += seasonList.loc[seasonList['PLAYER'] == playerName, 'GP'].sum()
    prumer = soucet.item()/pocet.item()
    print(f"Průměrný počet TO za posledních 10 sezon {playerName} je {prumer}")     
avgBLK_for_player_all_season('Michael Jordan')
        

# %% [markdown]
# # Kontrolní fce pro funkci avg_person_fauls()

# %%
def avgPF_for_player_all_season(playerName):
    soucet = 0
    pocet = 0
    for index in sheet_names:
        seasonList = dict_df.get(index)
        soucet += seasonList.loc[seasonList['PLAYER'] == playerName, 'PF'].sum()
        pocet += seasonList.loc[seasonList['PLAYER'] == playerName, 'GP'].sum()
    prumer = soucet.item()/pocet.item()
    print(f"Průměrný počet TO za posledních 10 sezon {playerName} je {prumer}")     
avgPF_for_player_all_season('Michael Jordan')
        

# %% [markdown]
# # Kontrolní funkce pro funkci avg_plus_minus()

# %%
def avgPlus_for_player_all_season(playerName):
    soucet = 0
    pocet = 0
    for index in sheet_names:
        seasonList = dict_df.get(index)
        soucet += seasonList.loc[seasonList['PLAYER'] == playerName, '+/-'].sum()
        pocet += seasonList.loc[seasonList['PLAYER'] == playerName, 'GP'].sum()
    prumer = soucet.item()/pocet.item()
    print(f"Průměrný počet TO za posledních 10 sezon {playerName} je {prumer}")     
avgPlus_for_player_all_season('Michael Jordan')
        


